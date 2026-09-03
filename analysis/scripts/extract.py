#!/usr/bin/env python3
"""包裝工時記錄表解析：把 xlsx 的 122 個工作表攤平成一份 rows.json。

每個工作表 = 一個品項，欄位固定為
    包裝日期 / 起始時間 / 結束時間 / 工時 / 瓶數

兩個資料特性要處理：
  1. 包裝日期沒有年份。109 個有資料的工作表全部依 10月→8月 排序且零違反，
     因此以 10-12 月為 2025、1-8 月為 2026 還原。
  2. 瓶數欄偶爾是混包寫法（例：A200/B209），代表同一批次做了兩種規格，
     依 Jackie 哥確認採加總處理（409 瓶）。

用法： python3 extract.py <xlsx路徑> [輸出目錄]
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

HEADER = ["包裝日期", "起始時間", "結束時間", "工時", "瓶數"]
DATE_RE = re.compile(r"^\s*(\d{1,2})月(\d{1,2})日")


def parse_bottles(v):
    """瓶數欄 → (數值, 是否為混包)。混包寫法把所有數字加總。"""
    if v is None:
        return None, False
    if isinstance(v, (int, float)):
        return float(v), False
    nums = re.findall(r"\d+", str(v))
    if not nums:
        return None, False
    return float(sum(int(n) for n in nums)), True


def parse_date(v):
    """'10月2日' → ('2025-10-02', 10, 2)。10-12 月屬 2025，1-9 月屬 2026。"""
    m = DATE_RE.match(str(v)) if v is not None else None
    if not m:
        return None
    mo, day = int(m.group(1)), int(m.group(2))
    year = 2025 if mo >= 10 else 2026
    return f"{year}-{mo:02d}-{day:02d}", mo, day


def extract(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rows, issues = [], []
    empty_sheets, header_bad = [], []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        values = list(ws.iter_rows(values_only=True))
        if not values or values[0][0] is None:
            empty_sheets.append(sheet)
            continue
        header = [str(x) if x is not None else "" for x in values[0][:5]]
        if header != HEADER:
            header_bad.append({"sku": sheet, "header": header})

        for excel_row, raw in enumerate(values[1:], start=2):
            date_v, start, end, hours, bottles_v = (list(raw) + [None] * 5)[:5]
            if date_v is None and bottles_v is None:
                continue  # Excel 預留空列

            where = {"sku": sheet, "row": excel_row, "date": str(date_v)}
            parsed = parse_date(date_v)
            if parsed is None:
                issues.append({**where, "kind": "日期無法解析"})
                continue
            iso, mo, day = parsed

            bottles, mixed = parse_bottles(bottles_v)
            if bottles is None:
                issues.append({**where, "kind": "瓶數空白"})
                continue
            if not isinstance(hours, (int, float)):
                issues.append({**where, "kind": "工時空白"})
                continue
            if hours <= 0:
                issues.append({**where, "kind": "工時為零或負數", "value": hours})
                continue

            rows.append({
                "sku": sheet,
                "date": iso,
                "month": iso[:7],
                "start": str(start)[:5] if start else None,
                "end": str(end)[:5] if end else None,
                "h": round(float(hours), 4),   # 原檔有浮點雜訊，統一收斂
                "b": bottles,
                "mixed": mixed,
            })

    return rows, {
        "issues": issues,
        "empty_sheets": empty_sheets,
        "header_mismatch": header_bad,
        "sheets_total": len(wb.sheetnames),
    }


def main():
    if len(sys.argv) < 2:
        sys.exit("用法: python3 extract.py <xlsx路徑> [輸出目錄]")
    xlsx = Path(sys.argv[1])
    out_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).resolve().parent.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    rows, meta = extract(xlsx)
    (out_dir / "rows.json").write_text(
        json.dumps({"rows": rows, "meta": meta}, ensure_ascii=False), encoding="utf-8")

    skus = sorted({r["sku"] for r in rows})
    dates = sorted({r["date"] for r in rows})
    print(f"來源      {xlsx.name}")
    print(f"工作表    {meta['sheets_total']} 個（{len(meta['empty_sheets'])} 個無資料）")
    print(f"有效紀錄  {len(rows):,} 筆 / {len(skus)} 個品項")
    print(f"資料期間  {dates[0]} ~ {dates[-1]}（{len(dates)} 個作業日）")
    print(f"混包紀錄  {sum(1 for r in rows if r['mixed'])} 筆")
    print(f"資料異常  {len(meta['issues'])} 筆" + ("" if not meta["issues"] else " ← 請看 audit.py"))
    print(f"輸出      {out_dir / 'rows.json'}")


if __name__ == "__main__":
    main()
